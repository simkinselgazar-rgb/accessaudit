"""Export ACR report data to an XLSX workbook using openpyxl."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import TestResult, ConformanceLevel, ReviewMeta

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour-coded fills for conformance levels
# ---------------------------------------------------------------------------
FILL_SUPPORTS = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")       # green
FILL_PARTIAL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")         # yellow
FILL_DOES_NOT = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")        # red
FILL_NA = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")              # gray
FILL_NE = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")              # purple
FILL_HEADER = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_SUPPORTS = Font(bold=True, color="2E7D32")
FONT_PARTIAL = Font(bold=True, color="F57F17")
FONT_DOES_NOT = Font(bold=True, color="C62828")
FONT_NA = Font(bold=True, color="757575")
FONT_NE = Font(bold=True, color="7B1FA2")

_SEVERITY_FILLS: dict[str, PatternFill] = {
    "high": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "medium": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "low": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "info": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

from report.utils import (
    FPC_MAPPING,
    FPC_NAMES,
    ResultProxy as _ResultProxy,
    build_fpc_rows as _build_fpc_rows,
    conformance_str as _conformance_str,
    criterion_sort_key as _sort_key,
    wrap_results as _wrap_results,
)


def _apply_conformance_style(cell, value: str) -> None:
    """Apply colour-coded fill and font to a conformance cell."""
    style_map = {
        ConformanceLevel.SUPPORTS.value: (FILL_SUPPORTS, FONT_SUPPORTS),
        ConformanceLevel.PARTIALLY_SUPPORTS.value: (FILL_PARTIAL, FONT_PARTIAL),
        ConformanceLevel.DOES_NOT_SUPPORT.value: (FILL_DOES_NOT, FONT_DOES_NOT),
        ConformanceLevel.NOT_APPLICABLE.value: (FILL_NA, FONT_NA),
        ConformanceLevel.NOT_EVALUATED.value: (FILL_NE, FONT_NE),
    }
    fill, font = style_map.get(value, (None, None))
    if fill:
        cell.fill = fill
    if font:
        cell.font = font


def _write_header_row(ws, headers: list[str], row: int = 1) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _add_summary_sheet(wb: Workbook, results: list[TestResult], meta: ReviewMeta, client_mode: bool = False) -> None:
    ws = wb.active
    ws.title = "Summary"

    _write_header_row(ws, ["Property", "Value"])

    meta_dict = meta.to_dict()
    info_rows = [
        ("Product Name", meta_dict.get("product_name", "")),
        ("Product Description", meta_dict.get("product_description", "")),
        ("Company", meta_dict.get("company_name", "")),
        ("Contact Name", meta_dict.get("contact_name", "")),
        ("Contact Email", meta_dict.get("contact_email", "")),
        ("URL", meta_dict.get("source_url", "")),
        ("Report Standard", "WCAG " + meta_dict.get("wcag_version", "2.2")),
        ("Coverage Level", meta_dict.get("coverage_level", "AA")),
        ("Date", meta_dict.get("created_at", "")),
        ("Review ID", meta_dict.get("review_id", "")),
        ("Notes", meta_dict.get("notes", "")),
        ("Evaluation Methods", meta_dict.get("evaluation_methods", "")),
        ("", ""),
        ("SUMMARY", ""),
    ]

    total = len(results)
    supports = sum(1 for r in results if _conformance_str(r.conformance_level) == ConformanceLevel.SUPPORTS.value)
    partially = sum(1 for r in results if _conformance_str(r.conformance_level) == ConformanceLevel.PARTIALLY_SUPPORTS.value)
    does_not = sum(1 for r in results if _conformance_str(r.conformance_level) == ConformanceLevel.DOES_NOT_SUPPORT.value)
    na_count = sum(1 for r in results if _conformance_str(r.conformance_level) == ConformanceLevel.NOT_APPLICABLE.value)
    ne_count = sum(1 for r in results if _conformance_str(r.conformance_level) == ConformanceLevel.NOT_EVALUATED.value)
    total_findings = sum(len(r.findings) for r in results)
    confidences = [r.confidence for r in results if r.confidence > 0]
    avg_conf = round(sum(confidences) / len(confidences), 3) if confidences else 0.0

    stats_rows = [
        ("Total Criteria", total),
        ("Supports", supports),
        ("Partially Supports", partially),
        ("Does Not Support", does_not),
        ("Not Applicable", na_count),
        ("Not Evaluated", ne_count),
        ("Total Findings", total_findings),
    ]
    if not client_mode:
        stats_rows.append(("Average Confidence", f"{avg_conf:.1%}"))

    row_num = 2
    for label, value in info_rows:
        ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=str(value))
        for c in range(1, 3):
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        row_num += 1

    for label, value in stats_rows:
        c1 = ws.cell(row=row_num, column=1, value=label)
        c1.font = Font(bold=True)
        c2 = ws.cell(row=row_num, column=2, value=str(value))
        for c in (c1, c2):
            c.border = THIN_BORDER
        # Colour code the conformance stats
        if label == "Supports":
            c2.fill = FILL_SUPPORTS
        elif label == "Partially Supports":
            c2.fill = FILL_PARTIAL
        elif label == "Does Not Support":
            c2.fill = FILL_DOES_NOT
        elif label == "Not Applicable":
            c2.fill = FILL_NA
        elif label == "Not Evaluated":
            c2.fill = FILL_NE
        row_num += 1

    # Legal disclaimer
    row_num += 1
    disclaimer_text = (
        "Legal Disclaimer: This report is provided for informational purposes "
        "only and does not constitute legal advice. Accessibility conformance "
        "was evaluated based on automated and manual testing methods and may "
        "not reflect all potential accessibility barriers. No warranty, express "
        "or implied, is made regarding the completeness or accuracy of this "
        "report."
    )
    c1 = ws.cell(row=row_num, column=1, value="Legal Disclaimer")
    c1.font = Font(bold=True, italic=True)
    c2 = ws.cell(row=row_num, column=2, value=disclaimer_text)
    c2.alignment = WRAP_ALIGNMENT
    for c in (c1, c2):
        c.border = THIN_BORDER

    _auto_width(ws)


def _add_criteria_sheet(
    wb: Workbook,
    title: str,
    results: list[TestResult],
    level_filter: str,
    client_mode: bool = False,
) -> None:
    filtered = [r for r in results if r.level.upper().strip() == level_filter]
    filtered.sort(key=lambda r: _sort_key(r.criterion_id))
    if not filtered:
        return

    ws = wb.create_sheet(title=title)
    if client_mode:
        headers = ["Criteria", "Name", "Conformance Level",
                    "Findings", "Remarks and Explanations"]
    else:
        headers = ["Criteria", "Name", "Conformance Level", "Confidence",
                    "Findings", "Summary"]
    _write_header_row(ws, headers)

    for row_idx, r in enumerate(filtered, start=2):
        conf_val = _conformance_str(r.conformance_level)
        ws.cell(row=row_idx, column=1, value=r.criterion_id).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=r.criterion_name).border = THIN_BORDER

        conf_cell = ws.cell(row=row_idx, column=3, value=conf_val)
        conf_cell.border = THIN_BORDER
        _apply_conformance_style(conf_cell, conf_val)

        if client_mode:
            ws.cell(row=row_idx, column=4, value=len(r.findings)).border = THIN_BORDER
            summary_cell = ws.cell(row=row_idx, column=5, value=r.summary)
            summary_cell.border = THIN_BORDER
        else:
            ws.cell(row=row_idx, column=4, value=f"{r.confidence:.1%}").border = THIN_BORDER
            ws.cell(row=row_idx, column=5, value=len(r.findings)).border = THIN_BORDER
            summary_cell = ws.cell(row=row_idx, column=6, value=r.summary)
            summary_cell.border = THIN_BORDER
        summary_cell.alignment = WRAP_ALIGNMENT

    _auto_width(ws)


def _add_tt_sheet(wb: Workbook, results: list[TestResult]) -> None:
    tt_rows = []
    for r in results:
        for tt in r.tt_results:
            if isinstance(tt, dict):
                tt_id = tt.get("tt_id", "")
                tt_name = tt.get("name", "")
                tt_result = tt.get("result", "")
            else:
                tt_id = tt.tt_id if hasattr(tt, "tt_id") else ""
                tt_name = tt.name if hasattr(tt, "name") else ""
                tt_result = tt.result.value if hasattr(tt.result, "value") else str(tt.result)
            tt_rows.append({
                "criterion_id": r.criterion_id,
                "criterion_name": r.criterion_name,
                "tt_id": tt_id,
                "name": tt_name,
                "result": tt_result,
            })
    if not tt_rows:
        return

    tt_rows.sort(key=lambda x: x.get("tt_id", ""))
    ws = wb.create_sheet(title="TT Results")
    headers = ["TT ID", "Test Name", "Criteria", "Criteria Name", "Result"]
    _write_header_row(ws, headers)

    tt_result_fills = {
        "PASS": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "FAIL": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "DNA": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
        "NOT TESTED": PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
    }

    for row_idx, row_data in enumerate(tt_rows, start=2):
        ws.cell(row=row_idx, column=1, value=row_data["tt_id"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=row_data["name"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=row_data["criterion_id"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=row_data["criterion_name"]).border = THIN_BORDER
        result_cell = ws.cell(row=row_idx, column=5, value=row_data["result"])
        result_cell.border = THIN_BORDER
        result_fill = tt_result_fills.get(row_data["result"])
        if result_fill:
            result_cell.fill = result_fill

    _auto_width(ws)


def _add_findings_sheet(wb: Workbook, results: list[TestResult], client_mode: bool = False) -> None:
    all_findings = []
    for r in results:
        for f in r.findings:
            entry = f.to_dict() if hasattr(f, "to_dict") else dict(f)
            entry["criterion_id"] = r.criterion_id
            entry["criterion_name"] = r.criterion_name
            all_findings.append(entry)
    if not all_findings:
        return

    sev_order = {"high": 0, "medium": 1, "low": 2, "info": 3}
    all_findings.sort(key=lambda x: (sev_order.get(x.get("severity", "info"), 4), x.get("criterion_id", "")))

    ws = wb.create_sheet(title="Findings")
    if client_mode:
        headers = [
            "Criteria", "Severity", "Element", "Location", "Selector", "Issue",
            "Impact", "Recommendation", "Evidence",
        ]
    else:
        headers = [
            "Criteria", "Severity", "Element", "Location", "Selector", "Issue",
            "Impact", "Recommendation", "Source", "Evidence",
        ]
    _write_header_row(ws, headers)

    for row_idx, f in enumerate(all_findings, start=2):
        ws.cell(row=row_idx, column=1, value=f.get("criterion_id", "")).border = THIN_BORDER
        sev = f.get("severity", "info")
        sev_cell = ws.cell(row=row_idx, column=2, value=sev.upper())
        sev_cell.border = THIN_BORDER
        sev_fill = _SEVERITY_FILLS.get(sev)
        if sev_fill:
            sev_cell.fill = sev_fill

        ws.cell(row=row_idx, column=3, value=f.get("element", "")).border = THIN_BORDER
        loc_cell = ws.cell(row=row_idx, column=4, value=f.get("location", ""))
        loc_cell.border = THIN_BORDER
        loc_cell.alignment = WRAP_ALIGNMENT
        ws.cell(row=row_idx, column=5, value=f.get("css_selector", "")).border = THIN_BORDER

        issue_cell = ws.cell(row=row_idx, column=6, value=f.get("issue", ""))
        issue_cell.border = THIN_BORDER
        issue_cell.alignment = WRAP_ALIGNMENT

        impact_cell = ws.cell(row=row_idx, column=7, value=f.get("impact", ""))
        impact_cell.border = THIN_BORDER
        impact_cell.alignment = WRAP_ALIGNMENT

        rec_cell = ws.cell(row=row_idx, column=8, value=f.get("recommendation", ""))
        rec_cell.border = THIN_BORDER
        rec_cell.alignment = WRAP_ALIGNMENT

        if client_mode:
            ev_col = 9
        else:
            ws.cell(row=row_idx, column=9, value=f.get("source", "")).border = THIN_BORDER
            ev_col = 10
        # Evidence column points at the captured screenshot when one is
        # available, so the operator (or client) can hyperlink straight
        # to the proof image when the workbook is opened.
        screenshot = f.get("screenshot_path", "") or ""
        ev_cell = ws.cell(row=row_idx, column=ev_col, value=screenshot or "")
        ev_cell.border = THIN_BORDER
        if screenshot:
            try:
                ev_cell.hyperlink = screenshot
                ev_cell.style = "Hyperlink"
            except Exception:
                # hyperlink decoration is non-critical; cell value already set
                pass

    _auto_width(ws)


def _add_fpc_sheet(wb: Workbook, results: list[TestResult]) -> None:
    fpc_rows = _build_fpc_rows(results)
    if not fpc_rows:
        return

    ws = wb.create_sheet(title="Functional Performance")
    headers = ["Code", "Criteria", "Conformance Level", "Remarks and Explanations"]
    _write_header_row(ws, headers)

    for row_idx, item in enumerate(fpc_rows, start=2):
        ws.cell(row=row_idx, column=1, value=item["code"]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=item["name"]).border = THIN_BORDER

        conf_cell = ws.cell(row=row_idx, column=3, value=item["conformance"])
        conf_cell.border = THIN_BORDER
        _apply_conformance_style(conf_cell, item["conformance"])

        remark_cell = ws.cell(row=row_idx, column=4, value=item.get("remark", ""))
        remark_cell.border = THIN_BORDER
        remark_cell.alignment = WRAP_ALIGNMENT

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_xlsx(
    results: list[TestResult],
    meta: ReviewMeta,
    output_path: str | Path,
    client_mode: bool = False,
) -> str:
    """Generate an XLSX accessibility conformance report.

    Sheets created:
    - **Summary** -- metadata and aggregate statistics
    - **Level A** -- WCAG Level A criteria results
    - **Level AA** -- WCAG Level AA criteria results
    - **Level AAA** -- WCAG Level AAA criteria results (only when AAA criteria exist)
    - **TT Results** -- Trusted Tester sub-test results
    - **Findings** -- all individual findings
    - **FPC** -- Section 508 Functional Performance Criteria (508 format only)

    Parameters
    ----------
    results : list[TestResult]
        Evaluated criteria results (may also be plain dicts).
    meta : ReviewMeta
        Review metadata.
    output_path : str | Path
        Destination file path for the .xlsx file.

    Returns
    -------
    str
        Absolute path of the written file.
    """
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    report_format = getattr(meta, "report_format", "508").lower()

    # Normalize: accept both TestResult objects and plain dicts
    results = _wrap_results(results)

    wb = Workbook()

    _add_summary_sheet(wb, results, meta, client_mode=client_mode)
    _add_criteria_sheet(wb, "Level A", results, "A", client_mode=client_mode)
    _add_criteria_sheet(wb, "Level AA", results, "AA", client_mode=client_mode)
    # Conditionally add Level AAA sheet when AAA criteria exist
    aaa_results = [r for r in results if r.level.upper().strip() == "AAA"]
    if aaa_results:
        _add_criteria_sheet(wb, "Level AAA", results, "AAA", client_mode=client_mode)
    if not client_mode:
        _add_tt_sheet(wb, results)
    _add_findings_sheet(wb, results, client_mode=client_mode)

    if report_format == "508":
        _add_fpc_sheet(wb, results)

    wb.save(str(output_path))
    logger.info("XLSX report written to %s", output_path)
    return str(output_path)
